#!/usr/bin/env python3
"""
build.py — turns core_data.xlsx into the static site's data.

What it does
------------
1. Reads the three sheets: Information, Build, GVG_team.
2. Extracts every picture that is stored *inside a cell* on the Information
   sheet (Excel's "Place in Cell" images) and writes it to assets/.
3. Writes data.json, which is the only file the website loads.

Nothing about the site is hardcoded: add rows to Build or GVG_team, re-run
this script, and the website updates itself.

Usage
-----
    python build.py                      # normal build
    python build.py --no-images          # data.json only, keep assets as-is
    python build.py --excel other.xlsx   # different source workbook
"""

from __future__ import annotations

import argparse
import io
import json
import re
import sys
import unicodedata
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

try:
    from PIL import Image

    HAS_PILLOW = True
except ImportError:  # pragma: no cover - Pillow is optional
    HAS_PILLOW = False


# --------------------------------------------------------------------------
# Configuration
# --------------------------------------------------------------------------

ROOT = Path(__file__).resolve().parent
DEFAULT_EXCEL = ROOT / "core_data.xlsx"
DATA_JSON = ROOT / "data.json"
ASSETS = ROOT / "assets"

SHEET_INFORMATION = "Information"
SHEET_BUILD = "Build"
SHEET_GVG = "GVG_team"

# Where each kind of extracted picture lands, and how wide it may be.
IMAGE_TARGETS = {
    "card": (ASSETS / "animus", 512),
    "portrait": (ASSETS / "animus", 256),
    "matrix": (ASSETS / "icons" / "matrix", 128),
    "shell": (ASSETS / "shells", 192),
    "passive": (ASSETS / "icons" / "passives", 128),
    "element": (ASSETS / "icons" / "elements", 96),
}

WEBP_QUALITY = 82

# XML namespaces used by the .xlsx package.
NS_MAIN = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
NS_RICHDATA = "{http://schemas.microsoft.com/office/spreadsheetml/2017/richdata}"
NS_REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"

warnings: list[str] = []


def warn(message: str) -> None:
    warnings.append(message)
    print(f"  ! {message}")


# --------------------------------------------------------------------------
# Small helpers
# --------------------------------------------------------------------------


def slugify(value: str) -> str:
    """'Cachi [The Guard]' -> 'cachi-the-guard'. Safe for filenames and URLs."""
    value = unicodedata.normalize("NFKD", str(value))
    value = re.sub(r"[^\w\s-]", " ", value, flags=re.UNICODE)
    value = re.sub(r"[\s_-]+", "-", value.strip())
    return value.strip("-").lower() or "item"


def key(value) -> str:
    """Loose matching key so 'Moon Bunny' and 'moonbunny' resolve to the same row."""
    if value is None:
        return ""
    return re.sub(r"[^0-9a-z\u0e00-\u0e7f]+", "", str(value).strip().lower())


def clean(value) -> str:
    """Cell value -> trimmed string. Excel errors and blanks become ''."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    if text.startswith("#") and text.endswith("!"):  # #VALUE!, #REF! ...
        return ""
    return text


def split_list(value) -> list[str]:
    """'12,8,6' -> ['12', '8', '6']. Values themselves are never altered."""
    text = clean(value)
    if not text:
        return []
    return [part.strip() for part in text.split(",")]


def col_letter(ref: str) -> str:
    return re.match(r"([A-Z]+)", ref).group(1)


def col_index(letter: str) -> int:
    n = 0
    for ch in letter:
        n = n * 26 + (ord(ch) - 64)
    return n


def row_number(ref: str) -> int:
    return int(re.search(r"(\d+)", ref).group(1))


# --------------------------------------------------------------------------
# In-cell image extraction
# --------------------------------------------------------------------------


def read_incell_images(xlsx_path: Path, sheet_name: str) -> dict[str, str]:
    """
    Map cell reference -> path inside the xlsx package, for pictures placed
    *in* cells on the given sheet.

    Excel stores these as "rich values". The chain is:
        cell[@vm]  ->  metadata.xml  ->  rdrichvalue.xml
                   ->  richValueRel.xml  ->  .rels  ->  xl/media/imageN.png
    """
    result: dict[str, str] = {}
    with zipfile.ZipFile(xlsx_path) as zf:
        names = set(zf.namelist())
        needed = {
            "xl/metadata.xml",
            "xl/richData/rdrichvalue.xml",
            "xl/richData/richValueRel.xml",
            "xl/richData/_rels/richValueRel.xml.rels",
        }
        if not needed.issubset(names):
            return result  # workbook simply has no in-cell pictures

        # vm index (1-based) -> rich value index
        meta = ET.fromstring(zf.read("xl/metadata.xml"))
        vm_to_rv: dict[int, int] = {}
        for future in meta.iter(NS_MAIN + "futureMetadata"):
            if future.get("name") != "XLRICHVALUE":
                continue
            for i, bk in enumerate(future.findall(NS_MAIN + "bk"), start=1):
                for rvb in bk.iter(NS_RICHDATA + "rvb"):
                    vm_to_rv[i] = int(rvb.get("i"))

        # rich value index -> local image id
        rvdata = ET.fromstring(zf.read("xl/richData/rdrichvalue.xml"))
        rv_to_image: dict[int, int] = {}
        for i, rv in enumerate(rvdata.findall(NS_RICHDATA + "rv")):
            values = rv.findall(NS_RICHDATA + "v")
            if values and values[0].text is not None:
                rv_to_image[i] = int(values[0].text)

        # local image id -> relationship id -> media file
        relmap = ET.fromstring(zf.read("xl/richData/richValueRel.xml"))
        rel_ids = [node.get(NS_REL + "id") for node in relmap]
        rels = ET.fromstring(zf.read("xl/richData/_rels/richValueRel.xml.rels"))
        rid_to_target = {node.get("Id"): node.get("Target") for node in rels}

        # find the sheet's xml part
        book = ET.fromstring(zf.read("xl/workbook.xml"))
        book_rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rid_to_sheet = {n.get("Id"): n.get("Target") for n in book_rels}
        sheet_part = None
        for sheet in book.iter(NS_MAIN + "sheet"):
            if sheet.get("name") == sheet_name:
                target = rid_to_sheet[sheet.get(NS_REL + "id")]
                sheet_part = "xl/" + target.lstrip("/").replace("xl/", "", 1)
                break
        if sheet_part is None or sheet_part not in names:
            return result

        sheet_xml = ET.fromstring(zf.read(sheet_part))
        for cell in sheet_xml.iter(NS_MAIN + "c"):
            vm = cell.get("vm")
            if not vm:
                continue
            try:
                media = rid_to_target[rel_ids[rv_to_image[vm_to_rv[int(vm)]]]]
            except (KeyError, IndexError):
                continue
            result[cell.get("r")] = "xl/" + media.lstrip("./").replace("../", "")
    return result


def asset_path(kind: str, name: str) -> Path:
    folder, _ = IMAGE_TARGETS[kind]
    suffix = "-card" if kind == "card" else ("-portrait" if kind == "portrait" else "")
    return folder / f"{slugify(name)}{suffix}"


def existing_asset(kind: str, name: str) -> str | None:
    """
    Reuse a picture that is already in assets/.

    This covers two cases: artwork you dropped in by hand, and artwork that was
    extracted by an earlier build but is no longer inside the workbook (saving
    an .xlsx with some tools strips in-cell pictures).
    """
    stem = asset_path(kind, name)
    for extension in (".webp", ".png", ".jpg", ".jpeg"):
        candidate = stem.with_suffix(extension)
        if candidate.exists():
            return str(candidate.relative_to(ROOT)).replace("\\", "/")
    return None


def save_image(
    xlsx_path: Path, member: str, kind: str, name: str, written: dict
) -> str | None:
    """Write one picture into assets/ and return its site-relative path."""
    folder, max_width = IMAGE_TARGETS[kind]
    folder.mkdir(parents=True, exist_ok=True)
    stem = asset_path(kind, name).name

    with zipfile.ZipFile(xlsx_path) as zf:
        try:
            raw = zf.read(member)
        except KeyError:
            warn(f"picture missing inside workbook for {kind} '{name}'")
            return None

    if HAS_PILLOW:
        try:
            image = Image.open(io.BytesIO(raw))
            image.load()
            if image.width > max_width:
                height = round(image.height * max_width / image.width)
                image = image.resize((max_width, height), Image.LANCZOS)
            if image.mode not in ("RGB", "RGBA"):
                image = image.convert("RGBA")
            out = folder / f"{stem}.webp"
            image.save(out, "WEBP", quality=WEBP_QUALITY, method=6)
            written[str(out)] = True
            return str(out.relative_to(ROOT)).replace("\\", "/")
        except Exception as exc:  # fall through to a plain copy
            warn(f"could not convert picture for '{name}' ({exc}); copying original")

    out = folder / f"{stem}{Path(member).suffix or '.png'}"
    out.write_bytes(raw)
    written[str(out)] = True
    return str(out.relative_to(ROOT)).replace("\\", "/")


# --------------------------------------------------------------------------
# Sheet readers
# --------------------------------------------------------------------------


def sheet_rows(worksheet) -> list[dict]:
    """Read a header-row sheet into a list of dicts keyed by column name."""
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean(h) for h in rows[0]]
    records = []
    for raw in rows[1:]:
        record = {}
        empty = True
        for header, value in zip(headers, raw):
            if not header:
                continue
            text = clean(value)
            record[header] = text
            if text:
                empty = False
        if not empty:
            records.append(record)
    return records


def read_information(worksheet, images: dict[str, str], xlsx_path: Path):
    """
    The Information sheet is a set of side-by-side lookup blocks:

        Animus | Profile | Card |  | Matrix |  | Full |  | Shell |  | ...

    Blocks are located by their header text, so columns can move without
    breaking the build.
    """
    # header text -> column letter
    headers: dict[str, str] = {}
    for cell in worksheet[1]:
        text = clean(cell.value)
        if text:
            headers[text.lower()] = cell.column_letter

    # column letter -> {row: value}
    values: dict[str, dict[int, str]] = {}
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            text = clean(cell.value)
            if text:
                values.setdefault(cell.column_letter, {})[cell.row] = text

    image_cols: dict[str, dict[int, str]] = {}
    for ref, member in images.items():
        if row_number(ref) >= 2:
            image_cols.setdefault(col_letter(ref), {})[row_number(ref)] = member

    def image_column_after(letter: str, skip: set[str]) -> str | None:
        start = col_index(letter)
        candidates = [
            c for c in image_cols if col_index(c) > start and c not in skip
        ]
        return min(candidates, key=col_index) if candidates else None

    written: dict[str, bool] = {}
    used_image_cols: set[str] = set()

    def build_lookup(header: str, kind: str, extra: dict | None = None) -> dict:
        letter = headers.get(header.lower())
        if not letter:
            warn(f"Information sheet has no '{header}' column; skipping that block")
            return {}
        icon_col = image_column_after(letter, used_image_cols)
        if icon_col:
            used_image_cols.add(icon_col)
        table = {}
        for row, name in sorted(values.get(letter, {}).items()):
            entry = {"name": name, "id": slugify(name)}
            member = image_cols.get(icon_col, {}).get(row) if icon_col else None
            path = (
                save_image(xlsx_path, member, kind, name, written)
                if member
                else existing_asset(kind, name)
            )
            if path:
                entry["icon"] = path
            for field, source_header in (extra or {}).items():
                source_col = headers.get(source_header.lower())
                if source_col:
                    entry[field] = values.get(source_col, {}).get(row, "")
            table[key(name)] = entry
        return table

    # --- Animus block (two pictures per row: Profile and Card) -------------
    animus_col = headers.get("animus")
    animus_art: dict[str, dict] = {}
    if animus_col:
        profile_col = headers.get("profile") or image_column_after(
            animus_col, used_image_cols
        )
        if profile_col:
            used_image_cols.add(profile_col)
        card_col = headers.get("card") or image_column_after(
            profile_col or animus_col, used_image_cols
        )
        if card_col:
            used_image_cols.add(card_col)
        for row, name in sorted(values.get(animus_col, {}).items()):
            entry = {"name": name, "id": slugify(name)}
            member = image_cols.get(profile_col, {}).get(row) if profile_col else None
            path = (
                save_image(xlsx_path, member, "portrait", name, written)
                if member
                else existing_asset("portrait", name)
            )
            if path:
                entry["portrait"] = path
            member = image_cols.get(card_col, {}).get(row) if card_col else None
            path = (
                save_image(xlsx_path, member, "card", name, written)
                if member
                else existing_asset("card", name)
            )
            if path:
                entry["card"] = path
            animus_art[key(name)] = entry
    else:
        warn("Information sheet has no 'Animus' column; artwork will be unavailable")

    matrices = build_lookup("Matrix", "matrix", extra={"full": "Full"})
    shells = build_lookup("Shell", "shell")
    passives = build_lookup("Shell Passive", "passive")
    elements = build_lookup("Element", "element")

    return animus_art, matrices, shells, passives, elements, written


# --------------------------------------------------------------------------
# Assembling data.json
# --------------------------------------------------------------------------


def make_build(record: dict, matrices: dict, shells: dict, passives: dict) -> dict:
    """One row of the Build sheet -> one build option."""

    def matrix_entry(name: str, fill: str) -> dict | None:
        if not name:
            return None
        item = {"name": name, "fill": fill}
        ref = matrices.get(key(name))
        if ref:
            item["id"] = ref["id"]
            if ref.get("icon"):
                item["icon"] = ref["icon"]
            if ref.get("full"):
                item["full"] = ref["full"]
        elif name:
            warn(f"matrix '{name}' is not listed on the Information sheet")
        return item

    fills = split_list(record.get("Matrix_Fill"))
    matrix_names = [
        record.get("Matrix_1", ""),
        record.get("Matrix_2", ""),
        record.get("Matrix_3", ""),
    ]
    matrix_list = []
    for i, name in enumerate(matrix_names):
        entry = matrix_entry(clean(name), fills[i] if i < len(fills) else "")
        if entry:
            matrix_list.append(entry)

    passive_list = []
    for column in ("Shell_Passive_1", "Shell_Passive_2", "Shell_Passive_3"):
        name = clean(record.get(column))
        if not name:
            continue
        item = {"name": name}
        ref = passives.get(key(name))
        if ref:
            item["id"] = ref["id"]
            if ref.get("icon"):
                item["icon"] = ref["icon"]
        else:
            warn(f"shell passive '{name}' is not listed on the Information sheet")
        passive_list.append(item)

    shell_name = clean(record.get("Shell"))
    shell = None
    if shell_name:
        shell = {"name": shell_name}
        ref = shells.get(key(shell_name))
        if ref:
            shell["id"] = ref["id"]
            if ref.get("icon"):
                shell["icon"] = ref["icon"]
        else:
            warn(f"shell '{shell_name}' is not listed on the Information sheet")

    return {
        "option": clean(record.get("Option")) or "1",
        "element": clean(record.get("Element")),
        "skill": clean(record.get("Skill")),
        "skillParts": split_list(record.get("Skill")),
        "matrixFill": clean(record.get("Matrix_Fill")),
        "matrices": matrix_list,
        "shell": shell,
        "passives": passive_list,
        "majorStat": clean(record.get("Major Stat")),
        "majorStats": split_list(record.get("Major Stat")),
        "minorStat": clean(record.get("Minor Stat")),
        "minorStats": split_list(record.get("Minor Stat")),
        "remark": clean(record.get("Remark")),
    }


def option_sort_key(build: dict):
    text = build.get("option", "")
    number = re.search(r"\d+", text)
    return (0, int(number.group())) if number else (1, text)


def build_payload(xlsx_path: Path, extract_images: bool) -> dict:
    print(f"Reading {xlsx_path.name}")
    workbook = load_workbook(xlsx_path, data_only=True)

    for name in (SHEET_INFORMATION, SHEET_BUILD, SHEET_GVG):
        if name not in workbook.sheetnames:
            print(f"ERROR: sheet '{name}' is missing from {xlsx_path.name}")
            sys.exit(1)

    images = read_incell_images(xlsx_path, SHEET_INFORMATION) if extract_images else {}
    if extract_images:
        print(f"  found {len(images)} in-cell pictures on {SHEET_INFORMATION}")
        if not images:
            warn(
                "no in-cell pictures found — the site will use whatever is already "
                "in assets/. Some tools strip pictures when they re-save an .xlsx; "
                "edit core_data.xlsx in Excel to keep them."
            )

    animus_art, matrices, shells, passives, elements, written = read_information(
        workbook[SHEET_INFORMATION], images, xlsx_path
    )
    if extract_images:
        print(f"  wrote {len(written)} files into assets/")

    # ---- Build sheet ------------------------------------------------------
    build_rows = sheet_rows(workbook[SHEET_BUILD])
    print(f"  {SHEET_BUILD}: {len(build_rows)} rows")

    by_animus: dict[str, dict] = {}
    order: list[str] = []
    for record in build_rows:
        name = clean(record.get("Animus"))
        if not name:
            warn("a Build row has no Animus name and was skipped")
            continue
        k = key(name)
        if k not in by_animus:
            art = animus_art.get(k, {})
            if not art:
                warn(f"'{name}' has builds but is not on the Information sheet")
            by_animus[k] = {
                "id": art.get("id") or slugify(name),
                "name": art.get("name") or name,
                "element": "",
                "portrait": art.get("portrait", ""),
                "card": art.get("card", ""),
                "builds": [],
            }
            order.append(k)
        by_animus[k]["builds"].append(make_build(record, matrices, shells, passives))

    # Animus listed on Information but without any build row yet.
    for k, art in animus_art.items():
        if k not in by_animus:
            by_animus[k] = {
                "id": art["id"],
                "name": art["name"],
                "element": "",
                "portrait": art.get("portrait", ""),
                "card": art.get("card", ""),
                "builds": [],
            }
            order.append(k)

    animus_list = []
    for k in order:
        entry = by_animus[k]
        entry["builds"].sort(key=option_sort_key)
        # Element comes from the builds; blank when there is no build yet.
        for build in entry["builds"]:
            if build["element"]:
                entry["element"] = build["element"]
                break
        element_ref = elements.get(key(entry["element"]))
        if entry["element"] and not element_ref:
            warn(f"element '{entry['element']}' is not on the Information sheet")
        entry["elementIcon"] = element_ref.get("icon", "") if element_ref else ""
        entry["buildCount"] = len(entry["builds"])
        animus_list.append(entry)

    animus_list.sort(key=lambda a: (a["buildCount"] == 0, a["name"].lower()))

    # ---- GVG_team sheet ---------------------------------------------------
    gvg_rows = sheet_rows(workbook[SHEET_GVG])
    print(f"  {SHEET_GVG}: {len(gvg_rows)} rows")

    animus_by_key = {key(a["name"]): a for a in animus_list}
    teams = []
    for index, record in enumerate(gvg_rows, start=1):
        team_name = clean(record.get("Team")) or f"Team {index}"
        slots = []
        # Slot prefixes are discovered from the headers, so a 4th slot ("D_")
        # would be picked up automatically.
        prefixes = sorted(
            {
                h.split("_")[0]
                for h in record
                if re.fullmatch(r"[A-Z]_animus", h, flags=re.IGNORECASE)
            }
        )
        for prefix in prefixes:
            name = clean(record.get(f"{prefix}_animus"))
            if not name:
                continue
            ref = animus_by_key.get(key(name))
            if not ref:
                warn(f"team '{team_name}' uses unknown Animus '{name}'")
            slot = {
                "name": ref["name"] if ref else name,
                "animusId": ref["id"] if ref else "",
                "portrait": ref.get("portrait", "") if ref else "",
                "card": ref.get("card", "") if ref else "",
                "element": ref.get("element", "") if ref else "",
                "matrices": [],
            }
            for n in (1, 2, 3):
                matrix_name = clean(record.get(f"{prefix}_Matrix_{n}"))
                if not matrix_name:
                    continue
                item = {"name": matrix_name}
                mref = matrices.get(key(matrix_name))
                if mref and mref.get("icon"):
                    item["icon"] = mref["icon"]
                slot["matrices"].append(item)
            shell_name = clean(record.get(f"{prefix}_Shell"))
            if shell_name:
                sref = shells.get(key(shell_name))
                slot["shell"] = {
                    "name": shell_name,
                    "icon": sref.get("icon", "") if sref else "",
                }
            slots.append(slot)
        teams.append({"id": slugify(team_name), "name": team_name, "slots": slots})

    def strip_key(table: dict) -> dict:
        return {entry["id"]: entry for entry in table.values()}

    return {
        "meta": {
            "source": xlsx_path.name,
            "generated": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
            "animusWithBuilds": sum(1 for a in animus_list if a["buildCount"]),
            "animusTotal": len(animus_list),
            "buildCount": len(build_rows),
            "teamCount": len(teams),
            "warnings": warnings,
        },
        "elements": strip_key(elements),
        "matrices": strip_key(matrices),
        "shells": strip_key(shells),
        "passives": strip_key(passives),
        "animus": animus_list,
        "teams": teams,
    }


# --------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(description="Build data.json from core_data.xlsx")
    parser.add_argument("--excel", default=str(DEFAULT_EXCEL), help="source workbook")
    parser.add_argument("--out", default=str(DATA_JSON), help="output json path")
    parser.add_argument(
        "--no-images",
        action="store_true",
        help="skip picture extraction (faster; leaves assets/ untouched)",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.excel).resolve()
    if not xlsx_path.exists():
        print(f"ERROR: {xlsx_path} not found")
        sys.exit(1)
    if not HAS_PILLOW and not args.no_images:
        print("  ! Pillow is not installed; pictures will be copied as PNG")

    payload = build_payload(xlsx_path, extract_images=not args.no_images)

    out_path = Path(args.out)
    out_path.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )

    meta = payload["meta"]
    size_kb = out_path.stat().st_size / 1024
    print(
        f"Wrote {out_path.name} ({size_kb:.0f} KB) — "
        f"{meta['animusWithBuilds']} Animus with builds "
        f"of {meta['animusTotal']}, {meta['buildCount']} build rows, "
        f"{meta['teamCount']} GVG teams"
    )
    if warnings:
        print(f"Finished with {len(warnings)} warning(s) — see the list above.")


if __name__ == "__main__":
    main()
